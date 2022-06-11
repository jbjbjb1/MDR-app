import logging
import os
import shutil
import sys
import time
import webbrowser
from contextlib import ExitStack

#import pkg_resources.py2_warn  # required for mdr-app to build in pyinstaller          # TODO only a temp fix to comment out
import win32com.client
import wx
from mailmerge import MailMerge
from openpyxl import load_workbook
from PyPDF3 import PdfFileMerger, PdfFileReader, PdfFileWriter

#Include logging for program at warning level or above
#logging.basicConfig(filename='mdr-app.log', filemode='a', level=logging.WARNING, format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')


class GenerateMDR():
    """ Tempoary holder for mdr generater program. """
    def __init__(self, project):
        

        def load_mdr_properties():
            """Load all properties from doc_properties.xlsx"""
            filename = project + r'\doc_properties.xlsx'
            # Load excel
            wb = load_workbook(filename=filename, read_only=True)
            mdr_properties = {}
            ws = wb['Properties']
            firstrow = True
            for row in ws.rows:
                # Don't import first row
                if firstrow == True:
                    firstrow = False
                    continue
                # But import all other rows
                else:
                    # Get variables out
                    temp_row = []
                    for cell in row:
                        temp_row.append(str(cell.value))
                    # Save variables to dictionary
                    mdr_properties[temp_row[0]] = temp_row[1]

            return mdr_properties


        def load_table_of_contents():
            """Load table of contents from doc_properties.xlsx"""
            filename = project + r'\doc_properties.xlsx'
            # Load excel
            wb = load_workbook(filename=filename, read_only=True)
            mdr_properties = []
            ws = wb['TableOfContents']
            firstrow = True
            for row in ws.rows:
                # Don't import first row
                if firstrow == True:
                    firstrow = False
                    continue
                # But import all other rows
                else:
                    # Get variables out
                    temp_row = {}
                    columns = ['Section', 'SectionDescription', 'PageNo']
                    for idx, cell in enumerate(row):
                        temp_row[columns[idx]] = str(cell.value)
                    # Save variables to dictionary
                    mdr_properties.append(temp_row)
            return mdr_properties


        def open_replace_save_doc(template_path, save_path, fields=None, table=None):
            """Opens a .docx file, replaces mergfields with variables, then save"""

            with MailMerge(template_path) as document:
                # Get list of fields to be merged
                #doc_merge_fields = document.get_merge_fields()
                # Merge fields as kwargs
                if fields:
                    document.merge(**fields)
                if table:
                    document.merge_rows('Section', table)  
                document.write(save_path)


        def convert_to_pdf(docs, source, destination):
            """Converts a list of docx files to pdf."""
            try:
                word = win32com.client.DispatchEx("Word.Application")
                for doc_name in docs:
                    doc = os.path.join(source, doc_name)
                    new_name = os.path.join(destination, doc_name)
                    new_name = new_name.replace(".docx", ".pdf")
                    worddoc = word.Documents.Open(doc)
                    worddoc.SaveAs(new_name, FileFormat = 17)
                    worddoc.Close()
            except Exception:
                    print('fn: convert docx to pdf', Exception)
            finally:
                    word.Quit()


        def count_pdf_pages(docs, source):
            """Count num pages of a pdf document, re-save in tempoary location."""
            num_of_sheets = 0
            for doc in docs:
                path = os.path.join(source, doc)
                pdf = PdfFileReader(open(path, 'rb'), strict=False)
                num_of_sheets += pdf.numPages
                if doc == 'contents_page.pdf':  # don't re-save if contents page
                    pass
                elif doc.split('.')[-1] != 'pdf':               # added so can have non-pdf docs in 
                    pass
                else:
                    pdfwrite = PdfFileWriter()                  # now just write a copy of the PDF to a temp location (fixes un-uniform docs?)
                    for page_count in range(pdf.numPages):      # loop over document and re-create
                        pages = pdf.getPage(page_count)
                        pdfwrite.addPage(pages)         
                    new_save_loc = ('\\').join(source.split('\\')[:-1]) + '\\output\\build\\resaved\\' + source.split('\\')[-1]   # save new files in this folder
                    if os.path.isdir(new_save_loc) == False:    # create dir if does not exist
                        os.makedirs(new_save_loc)
                    new_save_path = new_save_loc + '\\' + doc
                    with open(new_save_path, 'wb') as f:        # resave pdf
                        pdfwrite.write(f)
            return num_of_sheets


        def doc_pages():
            """Get a list of document pages for cover and contents page."""
            # Initiate variable
            page_nums = {'Cover Page': 1}
            page_num_contents = {'Cover Page': 1}
            # Create tempoary contents page
            template_path = template + r'\contents_page.docx'
            save_path = project + r'\output\build\temp\contents_page.docx'
            open_replace_save_doc(template_path, save_path, fields = mdr_properties, table = contents_properties)
            convert_to_pdf(['contents_page.docx'], file_path_build_temp, file_path_build_temp)
            # Count contents page
            value = count_pdf_pages(['contents_page.pdf'], file_path_build_temp)
            page_nums['Contents Page'] = value
            page_num_contents['Contents Page'] = value
            # Loop over sections
            for section in sections:
                page_count = 0
                # Default value
                page_nums['Section '+ section] = 1
                page_count += 1
                # Get the path of data files to be counted
                section_docs_path = os.path.join(file_path_data, section)
                # Get list of data files to be counted
                for parent, dirnames, filenames in os.walk(section_docs_path):
                    for filename in filenames:
                        if filename.split('.')[-1] != 'pdf':
                            pass
                        else:
                            value = count_pdf_pages([filename], section_docs_path)
                            page_nums['--file: ' + filename] = value
                            page_count += value
                page_num_contents['Section ' + section] = page_count
            # Add blank page at end
            page_nums['Blank Page'] = 1
            return page_nums, page_num_contents


        """ Build MDR function. """
        # Measure run time
        start_time = time.time()
        print('Starting MDR build...')

        # Define template location
        # Determine if running in Pyinstaller
        try:
            if getattr(sys, 'frozen') and hasattr(sys, '_MEIPASS'):
                print('running in a PyInstaller bundle')
                bundle_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
                template = os.path.join(bundle_dir, 'templates')
            else:
                print('running in a normal Python process')
                template = r'documents\templates'
        # TODO the below is getting triggered on a local VS code python run
        except AttributeError:
            print('***This program is running natively (not in compiled .exe).')
            template = r'documents\templates'

        file_path_output = os.path.join(project, 'output')
        file_path_build = os.path.join(project, 'output', 'build')
        file_path_build_temp = os.path.join(project, 'output', 'build', 'temp')
        file_path_build_pdf = os.path.join(project, 'output', 'build', 'pdf')
        file_path_build_resaved = os.path.join(project, 'output', 'build', 'resaved')
        file_path_dist = os.path.join(project, 'output', 'dist')
        file_path_data = os.path.join(project)

        # Delete output folder and re-build structure
        shutil.rmtree(file_path_output, ignore_errors=True)
        os.makedirs(file_path_build_pdf)      # makes all folders for leaf
        os.makedirs(file_path_build_temp)
        os.makedirs(file_path_dist)

        # Get MDR properties
        mdr_properties = load_mdr_properties()
        contents_properties = load_table_of_contents()
        sections = [x['Section'] for x in contents_properties]

        # Count number of pages
        page_nums, page_num_contents = doc_pages()

        # Generate 'Cover Page'
        template_path = template + r'\a_cover_page.docx'
        save_path = project + r'\output\build\a_cover_page.docx'
        mdr_properties['TotalPages'] = str(sum(page_nums.values()))         # add page number calculated
        open_replace_save_doc(template_path, save_path, fields = mdr_properties)

        # Generate 'Contents Page'
        template_path = template + r'\contents_page.docx'
        save_path = project + r'\output\build\contents_page.docx'
        cumulative_pages = page_num_contents['Cover Page'] +  page_num_contents['Contents Page']
        mdr_properties['pagenum'] = str(2)
        for entry in contents_properties:        # add page numbers calculated
            entry['PageNo'] = str(1 + cumulative_pages)
            cumulative_pages += page_num_contents['Section ' + entry['Section']]
        open_replace_save_doc(template_path, save_path, fields = mdr_properties, table = contents_properties)

        # Generate 'Blank Page'
        template_path = template + r'\blank_page.docx'
        save_path = project + r'\output\build\blank_page.docx'
        # pagenum in footer is taken care of by using "TotalPages of TotalPages" because it is the last page
        open_replace_save_doc(template_path, save_path, fields = mdr_properties, table = contents_properties)

        # Generate 'Section Break' pages
        for entry in contents_properties:
            template_path = template + r'\section_break.docx'
            save_path = project + r'\output\build\section_break_' + entry['Section'] + '.docx'
            fields = {**entry, **mdr_properties}    # merge properties
            fields['pagenum'] = entry['PageNo']     # changing variable to set pagenumber
            open_replace_save_doc(template_path, save_path, fields = fields)

        # Convert all Word files to PDF
        for parent, dirnames, filenames in os.walk(file_path_build):
            convert_to_pdf(filenames, file_path_build, file_path_build_pdf)
            break

        # Insert pdf files to make master pdf document
        merger = PdfFileMerger()

        # Insert cover and contents page
        with ExitStack() as stack:
            files = [stack.enter_context(open(os.path.join(file_path_build_pdf, fname), 'rb')) for fname in [
                'a_cover_page.pdf', 'contents_page.pdf']]
            [merger.append(file) for file in files]

        # Combine all sections with pages
        filenames = [os.path.join(file_path_build_pdf, 'section_break_' + section + '.pdf') for section in sections]
        with ExitStack() as stack:
            files = [stack.enter_context(open(fname, 'rb')) for fname in filenames]
            for file in files:
                # Add section page
                merger.append(file)
                # Get path to documents
                section = file.name.split('_')[-1].rstrip('.pdf')
                section_docs_path = os.path.join(file_path_build_resaved, section)         # this now combines re-saved docs
                # Loop over adding documents
                for parent, dirnames, filenames in os.walk(section_docs_path):
                    for filename in filenames:
                        if filename.split('.')[-1] != 'pdf':
                            filenames.remove(filename)
                    with ExitStack() as stack:
                        files = [stack.enter_context(open(
                            os.path.join(section_docs_path, fname), 'rb')) for fname in filenames]
                        [merger.append(file) for file in files]    

        # Add blank page at end
        with ExitStack() as stack:
            files = [stack.enter_context(open(os.path.join(file_path_build_pdf, fname), 'rb')) for fname in [
                    'blank_page.pdf']]
            [merger.append(file) for file in files]

        # Add bookmarks to PDF
        # TODO put description after section number in bookmark
        page_num = 0                            # page numbers are zero based
        for key, value in page_nums.items():
            merger.addBookmark(key, page_num)
            page_num += value                   # because page_num is only pages per section

        # Save output as single pdf
        final_filename = mdr_properties['DocumentNumber'] + '- REV' + mdr_properties['Rev1'] + '.pdf'
        filename = os.path.join(file_path_dist, final_filename)
        with open(filename, "wb") as output:
            merger.write(output)

        # Open result file
        os.startfile(filename)

        # Print run time
        print("Successful run in {:.1f} sec.".format(time.time() - start_time))

        
class MainWindow(wx.Frame):
    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, title=title)

        self.sizer = wx.BoxSizer(wx.VERTICAL)

        #Give some padding
        self.sizer.Add(250, 0)

        # Add buttons
        self.buttons = []
        button_list = ['Instructions', 'Generate DOC']
        for i, button_name in enumerate(button_list):
            self.buttons.append(wx.Button(self, -1, button_name, size=(-1, 50)))
            self.sizer.Add(self.buttons[i], 1, wx.EXPAND)

        #Layout sizers
        self.SetSizer(self.sizer)
        self.SetAutoLayout(1)
        self.sizer.Fit(self)
        self.Show()

        #Bind button events
        self.Bind(wx.EVT_BUTTON, self.OnInstructions, self.buttons[0])
        self.Bind(wx.EVT_BUTTON, self.OnGenerateMDR, self.buttons[1])

    def OnInstructions(self, event):
        # Open a website
        webbrowser.open('https://google.com')

    def OnGenerateMDR(self, event):
        # Open dialogue to build MDR
        dlg = wx.DirDialog (None, "Choose MDR directory", os.getcwd(),
                    wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST)
        dlg.ShowModal() 
        GenerateMDR(dlg.GetPath())      # project is the path of the folder selected
        dlg.Destroy()
        

# Start GUI
app = wx.App(False)
frame = MainWindow(None, "mdr-app")
app.MainLoop()